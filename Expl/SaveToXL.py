#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import openpyxl

def get_xl_letter(row_len, start_letter = u'A'):
    """
    :param row_len: Длина строки матрицы
    :param start_letter: Буква с которой начинается экспорт
    :return: Конечная буква выгрузки
    """
    try:
        start_letter = start_letter.upper()
        add_cols = ord(start_letter)
        add_cols-=64
    except (SyntaxError, AttributeError, TypeError):
        add_cols = 1
    if isinstance(row_len, int):
        final_len = row_len+add_cols
        l1 = chr(final_len%26+65)
        l2 = final_len/26
        if l2:
            return u"%s%s"%(chr(l2+64),l1)
        else:
            return l1
    else: return start_letter

def add_values_cells(data, cells):
    zips =  zip(data, cells)
    for row in zips:
        for v,c in zip(row[0], row[1]):
            c.value = v
class XlsIOError(Exception):
    def __init__(self, err_type, *args, **kwargs):
        self.err_type  = err_type
        super(XlsIOError, self).__init__(args, kwargs)

def exp_matrix(matrix, start_f, start_r, templ_path, save_as = False ):
    try:
        w_book = openpyxl.load_workbook(templ_path)
    except IOError:
        raise XlsIOError(1, templ_path)
    sheet = w_book.active
    max_letter = get_xl_letter(len(matrix[0]), start_f)
    cells_tmp =  tuple(sheet.iter_rows(u'%s%d:%s%d' % (start_f, start_r, max_letter,len(matrix)+start_r)))
    add_values_cells(matrix,cells_tmp)
    if save_as:
        try:
            w_book.save(save_as)
            os.system(u'start excel.exe %s' % save_as)
        except IOError:
            raise XlsIOError(2)
    else: return w_book

def exp_single_fa(fa_data, f22_ind, obj_name, expl_file, a_l, a_n, a_obj_l, a_obj_n, a_path, **kwargs):
    excel_path = os.path.dirname(expl_file)+ u'\\%s_xlsx_files' % os.path.basename(expl_file)[4:-4]
    if not os.path.exists(excel_path): os.makedirs(excel_path)
    dest_filename = u'%s\\%s.xlsx'%(excel_path, f22_ind)
    w_book = exp_matrix(fa_data, start_f = a_l, start_r = a_n, templ_path=a_path)
    sheet = w_book.active
    # sheet.title = u'Активный'
    sheet.cell(u'%s%s'%(a_obj_l, a_obj_n)).value = obj_name
    w_book.save(filename=dest_filename)
    os.system(u'start excel.exe %s' %  dest_filename)


if __name__ == u'__main__':
    print get_xl_letter(234, u'M')
    print get_xl_letter(24, u'A')