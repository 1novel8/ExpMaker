#!/usr/bin/env python
# -*- coding: utf-8 -*-

from __future__ import absolute_import
import os
import re
import openpyxl
# from openpyxl.reader.excel import load_workbook
# from openpyxl.styles import Border, Side




import shutil
# from openpyxl.worksheet.copier import WorksheetCopy
from time import gmtime, strftime

default_sheet_name = u'Expl_export'

class XlsError(Exception):
    def __init__(self, err_type, err_data = None, *args, **kwargs):
        self.err_type  = err_type
        self.err_data = err_data
        self.code = 404
        self.message = self.get_message_by_type(err_type)
        super(XlsError, self).__init__(self.message)


    def get_message_by_type(self, err_type):
        if err_type == 'no_dir':
            self.code = 40400
            return 'No work dir specified'
        if err_type == 'no_file':
            self.code = 40401
            return 'No work file specified'
        if err_type == 'no_template':
            self.code = 40402
            return 'failed to load template'
        if err_type == 'shutil_err':
            self.code = 40403
            return 'failed to copy from template'
        if err_type == 'failed_connection':
            self.code = 40404
            return 'Failed to connect to file'
        if err_type == 'failed_to_get_sheet':
            self.code = 40405
            return 'Failed to get sheet from workbook'
        if err_type == 'file_already_opened':
            self.code = 40406
            return u'Файл уже открыт для редактирования: \n%s' % self.err_data['filename']
        return 'Unexpected XLS error occurred'




class ExcelExporter(object):
    def __init__(self, work_xl_file_path, template_xl_file, default_path = None):
        ExcelExporter.test_xls_file(template_xl_file)
        self.default_path = default_path if default_path else os.path.dirname(template_xl_file)
        self.template_file = template_xl_file

        self.work_file = work_xl_file_path
        if not work_xl_file_path:
            return
        if not os.path.isfile(work_xl_file_path):
            return
        self.work_book = ExcelExporter.try_load_wb(work_xl_file_path)
        self.work_sheet = None

    def update_work_book(self):
        self.work_book = ExcelExporter.test_xls_file(self.work_file)

    def change_work_file(self, new_work_file):
        if self.work_file == new_work_file:
            print new_work_file
            print 'fail'
            return
        if os.path.isfile(new_work_file):
            os.remove(new_work_file)
        self.work_file = new_work_file
        ExcelExporter.copy_from_template(self.template_file, self.work_file)
        self.update_work_book()


    def run_table_export(self, tab_settings):
        print 'Excel runs', tab_settings.template_sheet_name
        self.update_sheet(tab_settings.template_sheet_name)
        self.work_sheet[tab_settings.date_out_cell] = strftime("%d.%m.%Y", gmtime())
        self.work_sheet[tab_settings.rayon_out_cell] = tab_settings.rayon_name
        self.work_sheet[tab_settings.shz_out_cell] = tab_settings.shz_name

        start_cell = tab_settings.start_export_cell
        for matrix_item in tab_settings.data_configure_list:
            if 'loaded_data' not in matrix_item or not(len(matrix_item['loaded_data'])):
                continue
            matrx = matrix_item['loaded_data']
            if 'header_title' in matrix_item and matrix_item['header_title']:
                self.work_sheet[start_cell] = matrix_item['header_title']
                end_merge_cell = ExcelExporter.get_cell_by_field_shift(start_cell, len(matrx[0]) - 1)
                self.merge_cells(start_cell, end_merge_cell)
                start_cell = ExcelExporter.get_cell_by_row_shift(start_cell, 1)

            if 'include_rowspan' in matrix_item and matrix_item['include_rowspan']:
                row_span = matrix_item['include_rowspan']
                if 'include_title' in matrix_item:
                    self.work_sheet[start_cell] = matrix_item['include_title']
                end_merge_cell = ExcelExporter.get_cell_by_field_shift(start_cell, row_span - 1)
                self.merge_cells(start_cell, end_merge_cell)

                shifted_start_cell = ExcelExporter.get_cell_by_field_shift(start_cell, row_span)
                self.export_matrix(matrx, shifted_start_cell)
                start_cell = ExcelExporter.get_cell_by_row_shift(start_cell, len(matrx))
            else:
                self.export_matrix(matrx, start_cell)
                start_cell = ExcelExporter.get_cell_by_row_shift(start_cell, len(matrx))
        print 'Excel running finished'
        try:
            self.work_book.save(filename = self.work_file)
            return 'OK'
        except Exception as err:
            raise XlsError('file_already_opened', {'filename': self.work_file})


    @staticmethod
    def get_cell_by_row_shift(basic_cell, shift_count):
        l, n = ExcelExporter.get_cell_let_num(basic_cell)
        int_n = int(n)
        int_n += shift_count
        return '%s%d' % (l, int_n)

    @staticmethod
    def get_cell_by_field_shift(basic_cell, shift_count):
        l, n = ExcelExporter.get_cell_let_num(basic_cell)
        l = ExcelExporter.get_xl_letter(shift_count - 1, l)
        return l + unicode(n)





    # Not supported yet
    # def upsert_sheet_from_template(self, sheet_name):
    #     templ_wb = ExcelExporter.test_xls_file(self.template_file)
    #     templ_sheet = ExcelExporter.get_sheet_by_name(templ_wb, sheet_name)
    #     ExcelExporter.remove_sheet_by_name(self.work_book, sheet_name)
    #     to_sheet = ExcelExporter.get_sheet_by_name(self.work_book, sheet_name)
    #     templ_sheet.parent = to_sheet.parent
    #     cp = WorksheetCopy(source_worksheet=templ_sheet, target_worksheet=to_sheet)
    #     cp.copy_worksheet()


    #
    # def set_bold_border(self, cell_range):
    #     rows = self.work_sheet.range(cell_range)
    #     border_type = Border.BORDER_THIN
    #     border_type = Border.BORDER_MEDIUM
    #
    #
    #     thin_border = Border(left=Side(style='thin'),
    #                          right=Side(style='thin'),
    #                          top=Side(style='thin'),
    #                          bottom=Side(style='thin'))
    #     self.work_sheet['U3'].border = thin_border
    #
    #     for row in rows:
    #         row[0].style.borders.left.border_style = border_type
    #         row[-1].style.borders.right.border_style = border_type
    #     for c in rows[0]:
    #         c.style.borders.top.border_style = border_type
    #     for c in rows[-1]:
    #         c.style.borders.bottom.border_style = border_type


    def merge_cells(self, start_cell, end_cell, unmerge = False):
        if not self.work_sheet:
            self.work_sheet = self.work_book.active
        try:

            if unmerge:
                self.work_sheet.unmerge_cells('%s:%s' % (start_cell, end_cell))
            else:
                self.work_sheet.merge_cells('%s:%s' % (start_cell, end_cell))

        except Exception as err:
            print 'already merged'


    @staticmethod
    def remove_sheet_by_name(w_book, sheet_name):
        if sheet_name in w_book.sheetnames:
            sh = w_book.get_sheet_by_name(sheet_name)
            w_book.remove(sh)


    def update_sheet(self, sheet_name):
        self.work_sheet = ExcelExporter.get_sheet_by_name(self.work_book, sheet_name)

    @staticmethod
    def get_sheet_by_name(work_book, sheet_name):
        sheet_name = unicode(sheet_name)
        if not sheet_name:
            raise XlsError('failed_to_get_sheet', {'sheet_name': sheet_name})
        if sheet_name not in work_book.sheetnames:
            work_book.create_sheet(title=sheet_name, index=0)
        return work_book.get_sheet_by_name(sheet_name)


    @staticmethod
    def is_xls_like(file_path):
        f_dir = os.path.dirname(file_path)
        f_base = os.path.basename(file_path)
        if not f_dir or not os.path.isdir(f_dir):
            return False
        if not f_base:
            return False
        if f_base.split('.')[-1] not in ['xls', 'xlsx']:
            return False
        return True


    def guarantee_work_xl_file(self):
        try:
            ExcelExporter.test_xls_file(self.work_file)
        except XlsError:
            if ExcelExporter.is_xls_like(self.work_file):
                try:
                    if os.path.isfile(self.work_file):
                        os.remove(self.work_file)
                except Exception as err:
                    print err
                    raise XlsError('failed_connection', {'file_path': self.work_file})
            else:
                try:

                    file_dir = os.path.dirname(self.work_file)
                    if not os.path.exists(file_dir):
                        os.makedirs(file_dir)

                    if not os.path.isdir(file_dir):
                        raise XlsError('no_dir', {'file_dir': file_dir})
                    self.work_file = os.path.join(file_dir, 'default.xlsx')
                except Exception as err:
                    print err
                    self.work_file = os.path.join(self.default_path, 'default.xlsx')
            ExcelExporter.copy_from_template(self.template_file, self.work_file)


    @staticmethod
    def try_load_wb(xls_path):
        try:
            return openpyxl.load_workbook(xls_path)
        except IOError:
            raise XlsError('failed_connection', {'file_path': xls_path})

    @staticmethod
    def copy_from_template(templ_file_path, destination_file):
        try:
            shutil.copyfile(templ_file_path, destination_file)
        except shutil.Error:
            raise XlsError('shutil_err', {'failed_file': destination_file})


    @staticmethod
    def test_xls_file(file_path):
        if not file_path or not os.path.isfile(file_path):
            raise XlsError('no_file', {file_path: file_path})
        return ExcelExporter.try_load_wb(file_path)

    @staticmethod
    def get_xl_letter(row_len, start_letter=u'A'):
        """
        :param row_len: Длина строки матрицы
        :param start_letter: Буква с которой начинается экспорт
        :return: Конечная буква выгрузки
        """
        try:
            start_letter = start_letter.upper()
            add_cols = ord(start_letter)
            add_cols -= 64
        except (SyntaxError, AttributeError, TypeError):
            add_cols = 1
        if isinstance(row_len, int):
            final_len = row_len + add_cols
            l1 = chr(final_len % 26 + 65)
            l2 = final_len / 26
            if l2:
                return u"%s%s" % (chr(l2 + 64), l1)
            else:
                return l1
        else:
            return start_letter

    def export_matrix(self, matrix, start_cell):
        try:
            if not self.work_sheet:
                self.work_sheet = self.work_book.active
            ExcelExporter.export_matrix_to_sheet(self.work_sheet, matrix, start_cell)
        except Exception as err:
            print err
            raise XlsError('matrix_export_failed', {'matrix': matrix})


    def set_sheet_cell(self, cel_letter, cel_num,  value):
        if not self.work_sheet:
            self.work_sheet = self.work_book.active
        self.work_sheet.cell(u'%s%s' % (cel_letter, cel_num)).value = value



    def start_os_file(self):
        try:
            os.system(u'start excel.exe %s' % self.work_file)
        except Exception as err:
            print err
            raise XlsError('no_file', {'file_path': self.work_file})


    @staticmethod
    def get_cell_let_num(cell_code):
        match = re.match(r"([a-z]+)([0-9]+)", cell_code, re.I)
        if match:
            cell =  match.groups()
            return cell[0], int(cell[1])
        else:
            return 'A', 1



    @staticmethod
    def export_matrix_to_sheet(xl_sheet, matrix, start_cell_name):
        start_f, start_r = ExcelExporter.get_cell_let_num(start_cell_name)
        for ind, row in enumerate(matrix):
            row_num = ind + start_r
            for f_ind, val in enumerate(row):
                cell_letter = ExcelExporter.get_xl_letter(f_ind - 1, start_f)
                xl_sheet[cell_letter + unicode(row_num)].value = val

        #
        # zips = zip(matrix, cells_tmp)
        # for row in zips:
        #     for v, c in zip(row[0], row[1]):
        #         c.value = v
