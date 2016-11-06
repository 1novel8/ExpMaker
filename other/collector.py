#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import openpyxl
from DbTools import DBConn
from dbfread import DBF

from os import listdir
from os import path
project_dir = os.getcwd()

def is_valid_xl_file(f):
    if not path.isfile(path.join(project_dir, f)):
        return False
    if not (f.lower()[-4:] == u'xlsx' or f.lower()[-3:] == u'xls'):
        return False
    return True

xls_files = [f for f in listdir(project_dir) if is_valid_xl_file(f)]

def run_searched_files(searched_files):
    for f in searched_files:
        if isinstance(f, str):
           get_w_sheets(f)


def get_wb(xls_file_path):
    return openpyxl.load_workbook(filename=xls_file_path, use_iterators=True)


def get_w_sheets(xls_file_path):
    wb = get_wb(xls_file_path)
    return wb.get_sheet_names()


default_sheet_name = u'sheet_1'
def get_sheet_by_name(work_book, sh_name = None):
    sh_name = unicode(sh_name)
    if sh_name in work_book.sheetnames:
        sheet_name = sh_name
    else:
        sheet_name = sh_name if sh_name else default_sheet_name
        work_book.create_sheet(title=sheet_name, index=0)
    return work_book.get_sheet_by_name(sheet_name)


def get_dbf_table(file_path):
    try:
        db = DBF(file_path)
    except Exception as err:
        print err
        raise err


    out_table = []
    for rec in db:
        new_rec = {}
        for key in rec:
            if key[0] != '_':
                new_rec[key] = rec[key]
        out_table.append(new_rec)

    return out_table


def getMdbStructureFromXls(xls_file_path):

    available_types = {
        'Long': 'INTEGER',
        'Floa': 'REAL',
        'Shor': 'SMALLINT',
        'Text': 'VARCHAR',

    }

    wb = get_wb(xls_file_path)
    shts = wb.get_sheet_names()
    structure_doc = {}
    for sh_name in shts:

        sheet = get_sheet_by_name(wb, sh_name)
        structure_doc[sh_name] = {
            'tab_name': sheet.cell(u'B1').value,
            'tab_fields': ['NPRIZNAK', 'CTITLE'],
            'dbf_fields': {
                'NPRIZNAK': 'NPRIZNAK',
                'CTITLE': 'CTITLE',
            },
            'mdb_types': {
                'NPRIZNAK': 'VARCHAR',
                'CTITLE': 'VARCHAR',
            }
        }

        f_cell_num = 3
        while True:
            try:
                field_name = sheet.cell(u'C%d' % f_cell_num).value
            except Exception:
                break
            if not field_name:
                break
            dbf_field = sheet.cell(u'D%d' % f_cell_num).value
            dbf_type = sheet.cell(u'E%d' % f_cell_num).value
            try:
                mdb_type = available_types[dbf_type[:4]]
            except KeyError:
                mdb_type = 'VARCHAR'
            except TypeError:
                print 'warning on table' + sh_name
                mdb_type = 'VARCHAR'
            structure_doc[sh_name]['tab_fields'].append(field_name)
            structure_doc[sh_name]['dbf_fields'][field_name] = dbf_field
            structure_doc[sh_name]['mdb_types'][field_name] = mdb_type
            f_cell_num += 1
    return structure_doc


def get_dbf_file_names(fls_path):

    def is_valid_dbf_file(f):
        if not path.isfile(path.join(fls_path, f)):
            return False
        if not f.lower()[-3:] == u'dbf':
            return False
        return True

    return [f for f in listdir(fls_path ) if is_valid_dbf_file(f)]


def get_table_type(dbf_name):
    tab_type = dbf_name.split('_')[0]
    if not tab_type:
        raise Exception('Can\'t identify table type')
    tab_type = tab_type.upper()
    if tab_type.split('0')[0] == 'T':
        tab_type = ''.join(tab_type.split('0'))
    return tab_type

def get_table_code(dbf_name):
    tab_code = dbf_name.split('_')[1]
    return tab_code.split('.')[0]



def handle_int(val):
    try:
        return int(val)
    except Exception:
        raise


def handle_float(val):
    try:
        return float(val)
    except Exception:
        raise Exception('Unsupported type')

def synchronize_by_type(val, type):
    check_types = {
        'INTEGER': lambda x: handle_int(x),
        'REAL': lambda x: handle_float(x),
        'SMALLINT': lambda x: handle_int(x),
        'VARCHAR': lambda x: unicode(x),
    }
    return check_types[type](val)


def collect_data_to_mdb(dbf_path, mdb_path, required_mdb_str):
    dbf_files_names = get_dbf_file_names(dbf_path)
    mdb_db = DBConn(mdb_path)
    mdb_out_structure = mdb_db.get_tabs_with_field_names()


    for tab_type in required_mdb_str:
        should_be_t_name = required_mdb_str[tab_type]['tab_name']
        should_be_t_fields = required_mdb_str[tab_type]['tab_fields']
        if should_be_t_name not in mdb_out_structure:
            raise Exception('table for %s is not found in out mdb database' % tab_type)
        for field in should_be_t_fields:
            required_f_type = required_mdb_str[tab_type]['mdb_types'][field]
            if field not in mdb_out_structure[should_be_t_name]:
                print(' adding field %s to table %s' % (field, should_be_t_name))
                # add new field
                mdb_db.add_field_if_not_exists(should_be_t_name, field, required_f_type)
            elif mdb_out_structure[should_be_t_name][field] != required_f_type:
                if required_f_type in ['SMALLINT', 'INTEGER']:
                    continue
                else:
                    raise Exception('%s field type for table %s doesnt match with requirement' % (field, should_be_t_name))

    del mdb_out_structure

    for dbf_name in dbf_files_names:
        tab_type = get_table_type(dbf_name)
        tab_code = get_table_code(dbf_name)
        if not required_mdb_str.has_key(tab_type):
            print 'Faild to specify dbf file %s with table in excel structure' % dbf_name
            continue
        if not tab_code:
            print 'There is no code for dbf file ' + dbf_name
            continue

        dbf_data = get_dbf_table(path.join(dbf_path, dbf_name))
        tab_name = required_mdb_str[tab_type]['tab_name']
        ordered_fields = required_mdb_str[tab_type]['tab_fields']
        dbf_fields = required_mdb_str[tab_type]['dbf_fields']
        mdb_types = required_mdb_str[tab_type]['mdb_types']
        one_time_flag = False
        for row in dbf_data:
            insert_values = []
            for f in ordered_fields:
                f_val = ''
                try:
                    f_val = row[dbf_fields[f]]
                except KeyError:
                    if not one_time_flag:
                        one_time_flag = True
                        print 'for dbf %s field %s was not found' % (dbf_name, f)
                #TODO: make fileds that are comes not from dbf
                if f == 'NR_User':
                    f_val = tab_code

                # synchronizing with type
                f_val = synchronize_by_type(f_val, mdb_types[f])
                insert_values.append(f_val)
            mdb_db.insert_row(tab_name, ordered_fields, insert_values)








if __name__ == u'__main__':
    xls_str_file = project_dir + '\\structure.xlsx'
    mdb_out_file = project_dir + '\\ProbaCO6258.mdb'
    dbf_fls_path = project_dir + '\\dbf_files'
    xls_str = getMdbStructureFromXls(xls_str_file)
    collect_data_to_mdb(dbf_fls_path, mdb_out_file, xls_str)

    #TODO: make a copy of mdb file
