#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
import openpyxl

default_sheet_name = u'Expl_export'

def get_xl_letter(row_len, start_letter = u'A'):
    """
    :param row_len: Длина строки матрицы
    :param start_letter: Буква с которой начинается экспорт
    :return: Конечная буква выгрузки
    """
    try:
        start_letter = start_letter.upper()
        add_cols = ord(start_letter)
        add_cols-=64
    except (SyntaxError, AttributeError, TypeError):
        add_cols = 1
    if isinstance(row_len, int):
        final_len = row_len+add_cols
        l1 = chr(final_len%26+65)
        l2 = final_len/26
        if l2:
            return u"%s%s"%(chr(l2+64),l1)
        else:
            return l1
    else: return start_letter

def add_values_cells(data, cells):
    zips =  zip(data, cells)
    for row in zips:
        for v,c in zip(row[0], row[1]):
            c.value = v
class XlsIOError(Exception):
    def __init__(self, err_type, file_name, *args, **kwargs):
        self.err_type  = err_type
        self.file_name = file_name
        super(XlsIOError, self).__init__(args, kwargs)

#TODO: переделать бы в класс XlExporter

def export_matrix_to_sheet(xl_sheet, matrix, start_f, start_r):
    max_letter = get_xl_letter(len(matrix[0]), start_f)
    cells_tmp =  tuple(xl_sheet[u'%s%d:%s%d' % (start_f, start_r, max_letter,len(matrix)+start_r)])
    add_values_cells(matrix,cells_tmp)

def try_load_wb(l_path):
    try:
        return openpyxl.load_workbook(l_path)
    except IOError:
        raise XlsIOError(1, l_path)

def get_sheet_by_name(work_book, sh_name = None):
    if not sh_name:
        return work_book.active
    sh_name = unicode(sh_name)
    if sh_name in work_book.sheetnames:
        sheet_name = sh_name
    else:
        sheet_name = sh_name
        work_book.create_sheet(title=sh_name, index=0)
    return work_book.get_sheet_by_name(sheet_name)

def exp_matrix(matrix, start_f, start_r, templ_path, save_as, is_xls_start, sh_name = None):
    """
    You can give templ_path parameter and save w_book or give worksheet parameter and export matrix without saving
    """
    if templ_path:
        w_book = try_load_wb(templ_path)
        sheet = get_sheet_by_name(w_book, sh_name)
    else:
        raise XlsIOError(4, templ_path)
    export_matrix_to_sheet(sheet, matrix, start_f, start_r)
    try:
        w_book.save(save_as)
        if is_xls_start:
            os.system(u'start excel.exe %s' % save_as)
    except IOError:
        raise XlsIOError(2, save_as)

def exp_single_fa(fa_data, f22_ind, obj_name, expl_file, a_l, a_n, a_obj_l, a_obj_n, a_path, a_sh_name, is_xls_start, **kwargs):
    excel_path = os.path.dirname(expl_file)+ u'\\%s_xlsx_files' % os.path.basename(expl_file)[4:-4]
    if not os.path.exists(excel_path):
        os.makedirs(excel_path)
    dest_filename = u'%s\\%s.xlsx'%(excel_path, f22_ind)
    w_book = try_load_wb(a_path)
    sheet = get_sheet_by_name(w_book, a_sh_name)
    sheet.cell(u'%s%s'%(a_obj_l, a_obj_n)).value = obj_name
    export_matrix_to_sheet(sheet, fa_data, a_l, a_n)
    if not os.path.isfile(dest_filename):
        w_book.save(filename=dest_filename)
        if is_xls_start:
            os.system(u'start excel.exe %s' % dest_filename)
    else:
        try:
            os.remove(dest_filename)
        except Exception as err:
            print err
            raise XlsIOError(1, dest_filename)
        w_book.save(filename=dest_filename)
        if is_xls_start:
            os.system(u'start excel.exe %s' % unicode(dest_filename))
        # raise XlsIOError(3,dest_filename)


if __name__ == u'__main__':
    print get_xl_letter(234, u'M')
    print get_xl_letter(24, u'A')