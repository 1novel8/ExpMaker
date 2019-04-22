#!/usr/bin/env python
# -*- coding: utf-8 -*-

import os
from openpyxl import load_workbook, utils
from core.errors import XlsError


class XlExporter:
    def __init__(self, out_filename=None, template_path=None):
        self.template_path = template_path
        self.out_filename = out_filename.replace(' ', '_').replace('/', '').replace('"', '')
        if not os.path.exists(os.path.dirname(self.out_filename)):
            os.makedirs(os.path.dirname(self.out_filename))

    def export_matrix(self, matrix, start_f, start_r, sh_name=None, **kwargs):
        """
        You can give templ_path parameter and save w_book or give worksheet parameter and export matrix without saving
        """
        w_book = self.try_load_wb(self.template_path)
        sheet = self.get_sheet_by_name(w_book, sh_name)
        self.export_matrix_to_sheet(sheet, matrix, start_f, start_r)
        try:
            w_book.save(self.out_filename)
        except IOError:
            raise XlsError('not_found', self.out_filename)

    def start_excel(self):
        os.system('start excel.exe %s' % self.out_filename)

    def exp_single_fa(self, fa_data, obj_name, a_l, a_n, a_obj_l, a_obj_n, a_path, a_sh_name, **kwargs):
        w_book = self.try_load_wb(a_path)
        sheet = self.get_sheet_by_name(w_book, a_sh_name)
        sheet.cell(a_obj_n, utils.column_index_from_string(a_obj_l)).value = obj_name
        self.export_matrix_to_sheet(sheet, fa_data, a_l, a_n)
        if not os.path.isfile(self.out_filename):
            w_book.save(filename=self.out_filename)
            if kwargs['is_xls_start']:
                self.start_excel()
        else:
            try:
                os.remove(self.out_filename)
            except Exception as err:
                print(err)
                raise XlsError('already_opened', self.out_filename)
            w_book.save(filename=self.out_filename)

    @staticmethod
    def get_xl_letter(row_len, start_letter='A'):
        """
        :param row_len: Длина строки матрицы
        :param start_letter: Буква с которой начинается экспорт
        :return: Конечная буква выгрузки
        """
        try:
            start_letter = start_letter.upper()
            add_cols = ord(start_letter)
            add_cols -= 64
        except (SyntaxError, AttributeError, TypeError):
            add_cols = 1
        if isinstance(row_len, int):
            final_len = row_len+add_cols
            l1 = chr(final_len % 26+65)
            l2 = final_len/26
            if l2:
                return "%s%s" % (chr(int(l2+64)), l1)
            else:
                return l1
        else:
            return start_letter

    @staticmethod
    def add_values_cells(data, cells):
        zips = zip(data, cells)
        for row in zips:
            for v, c in zip(row[0], row[1]):
                c.value = v

    @staticmethod
    def try_load_wb(l_path):
        if not os.path.isfile(l_path):
            raise XlsError('not_found', l_path)
        try:
            return load_workbook(filename=l_path, data_only=False)
        except IOError:
            raise XlsError('already_opened', l_path)
        except Exception:
            raise XlsError('load_failed', l_path)

    def export_matrix_to_sheet(self, xl_sheet, matrix, start_f, start_r):
        max_letter = self.get_xl_letter(len(matrix[0]), start_f)
        cells_tmp = tuple(xl_sheet['%s%d:%s%d' % (start_f, start_r, max_letter, len(matrix)+start_r)])
        self.add_values_cells(matrix, cells_tmp)

    @staticmethod
    def get_sheet_by_name(work_book, sh_name=None):
        if not sh_name:
            return work_book.active
        sh_name = str(sh_name)
        if sh_name in work_book.sheetnames:
            sheet_name = sh_name
        else:
            sheet_name = sh_name
            work_book.create_sheet(title=sh_name, index=0)
        return work_book.get_sheet_by_name(sheet_name)
